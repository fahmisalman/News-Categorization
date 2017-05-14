from __future__ import division
from openpyxl import Workbook

import Preprocessing
import glob
import re


def loadData(location):
    bag = []
    n = 0
    for name in glob.glob(location):
        file = open(name)
        s = file.read()
        s = s.lower()
        s = re.sub(r'[^a-z]', ' ', s)
        bag += s.split()
        n += 1
    pre = Preprocessing.Preprocessing()
    bag = pre.stopwordRemoval(bag)
    bag = pre.lemmatization(bag)
    return bag, n

def likelihoodProbabilities(list, bag):
    likelihood = []
    for i in range(len(list)):
        likelihood.append(bag.count(list[i]))
    for i in range(len(likelihood)):
        likelihood[i] += 1
    total = sum(likelihood)
    for i in range(len(likelihood)):
        likelihood[i] /= total
    return likelihood

def saveToExcel(ws, column_cell, list, string):
    ws[column_cell + str(1)] = string
    for row, i in enumerate(list):
        ws[column_cell + str(row + 2)] = str(i)


if __name__ == '__main__':
    sport, nSport = loadData('bbc-2/sport/*')
    business, nBusiness = loadData('bbc-2/business/*')
    politics, nPolitics = loadData('bbc-2/politics/*')
    entertainment, nEntertainment = loadData('bbc-2/entertainment/*')
    tech, nTech = loadData('bbc-2/tech/*')
    bag = sport + business + politics + entertainment + tech
    wordList = list(set(bag))
    likelihoodSport = likelihoodProbabilities(wordList, sport)
    likelihoodBusiness = likelihoodProbabilities(wordList, business)
    likelihoodPolitics = likelihoodProbabilities(wordList, politics)
    likelihoodEntertainment = likelihoodProbabilities(wordList, entertainment)
    likelihoodTech = likelihoodProbabilities(wordList, tech)

    wb = Workbook()
    ws = wb.active
    saveToExcel(ws, 'A', wordList, "wordlist")
    saveToExcel(ws, 'B', likelihoodSport, "sport")
    saveToExcel(ws, 'C', likelihoodBusiness, "business")
    saveToExcel(ws, 'D', likelihoodPolitics, "politics")
    saveToExcel(ws, 'E', likelihoodEntertainment, "entertainment")
    saveToExcel(ws, 'F', likelihoodTech, "tech")
    wb.save("sample.xlsx")
